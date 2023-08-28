import time
import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import sys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os.path
from tkinter import filedialog
from tkinter import *
import logging
from selenium.webdriver.remote.remote_connection import LOGGER


dict_role = {
    1:"Ученик",
    2:"Родитель",
    3:"Ученик и Родитель",
    4:"Учитель",
    5:"Сотрудники"
}
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
chrome_options.add_argument("--lang=ru")

def save_file(file_path):
    while file_path == '':
        root = Tk()
        root.attributes('-topmost', True)
        root.withdraw()
        file_path = filedialog.asksaveasfilename(parent=root, defaultextension='.xlsx')
        if file_path != '':
            return file_path

def main(admin,password,role,name,groupNumber,groups,language):
        roles = int(role)
        LOGGER.setLevel(logging.WARNING)
        driver = webdriver.Chrome(service=Service(executable_path=executable_path), options=chrome_options)
        driver.get("https://login.kundelik.kz/login")
        driver.maximize_window()
        status = False
        if groups == "false" or groups =="False":
            group = False
        else:
            group = True

        if createExcel(language) == 13:
            driver.quit()
            # raise ValueError(14)1
            sys.stdout.write("14")
            return

        WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Кіру']")))

        #login to site by admin
        driver.find_element(By.XPATH, "//input[@name='login']").send_keys(admin)
        driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//input[@value='Кіру' or @value='Войти']").click()
        ###



        #Edu->mySchool->People
        try:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@data-test-id = 'new-post-button']")))
        except:
            driver.quit()
            # raise ValueError(1)
            sys.stdout.write("1")

        getLanguage = driver.find_element(By.XPATH, "//div[@class='header-localization-select__info']/div[1]").text
        if language == "kz":
            if getLanguage != "Қаз":
                driver.get(
                    "https://kundelik.kz/api/localization/change?newCulture=kk-KZ&currentUrl=http%3a%2f%2fkundelik.kz%2fteachers")
        else:
            if getLanguage == "Қаз":

                driver.get(
                    "https://kundelik.kz/api/localization/change?newCulture=ru-RU&currentUrl=http%3a%2f%2fkundelik.kz%2fteachers")

            # raise ValueError(1) from None

        driver.find_element(By.XPATH, "//li/a[@title='Моя школа' or @title='Менің мектебім']").click()
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div/button[@data-test-id='send-post']")))

        driver.find_element(By.XPATH, "//div[@class='tabs']//a[text()='Люди' or text()='Адамдар']").click()
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Найти' or @value='Табу']")))


        if group:
            status = massiveGoToAll(driver, False,language)

        else:
            if len(name)>1:
                driver.find_element(By.XPATH, "//div/input[@id='search']").send_keys(name)
                time.sleep(0.5)
                if groupNumber != -1:
                    driver.find_element(By.XPATH, "//input[@id='class']").send_keys(groupNumber)

                time.sleep(0.5)
                driver.find_element(By.XPATH, "//input[@value='Найти' or @value='Табу']").click()
                foundStudent = ""
                try:
                    WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
                    foundStudent = driver.find_element(By.XPATH, "//p[@class='found']/strong").text
                except:
                    driver.quit()
                    # raise ValueError(9)
                    sys.stdout.write("9")
                if int(foundStudent) == 1:
                    time.sleep(0.5)
                    if driver.find_element(By.XPATH,"//p[@class='found']/strong"):
                        if roles == 1:
                            status = goToStudent(driver,False,language)
                        elif roles == 2:
                            status = gotoDirectParent(driver,language)
                        elif roles == 3:
                            status = goToStudent(driver,True,language)
                        elif roles == 4:
                            status = gotoTeacher(driver,language)
                        else:
                            pass
                else:
                    if roles == 1:
                        status = gotoAllStudent(driver, False,language)
                    elif roles == 2:
                        status = gotoAllStudentParent(driver,language)
                    elif roles == 3:
                        status = gotoAllStudent(driver, True,language)

            elif len(name)==0 and groupNumber != "":
                driver.find_element(By.XPATH, "//input[@id='class']").send_keys(groupNumber)
                driver.find_element(By.XPATH, "//input[@value='Найти' or @value='Табу']").click()
                try:
                    WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
                except:
                    driver.quit()
                    # raise ValueError(11)
                    sys.stdout.write("20")
                if roles == 1:
                    status =gotoAllStudent(driver, False,language)
                elif roles == 2:
                    status =gotoAllStudentParent(driver,language)
                elif roles == 3:
                    status =gotoAllStudent(driver, True,language)
                else:
                    status =gotoAllStudent(driver, False,language)


            elif roles==1:
                driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Ученики' or text()='Оқушылар']").click()
                status = gotoAllStudent(driver, False,language)
            elif roles ==2:
                driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Родители' or text()='Ата-аналар']").click()
                status = gotoAllStudentParent(driver,language)
            elif roles == 3:
                driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Ученики' or text()='Оқушылар']").click()
                status = gotoAllStudent(driver,True,language)
            elif roles == 5:
                driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Сотрудники' or text()='Қызметкерлер']").click()
                status = gotoAllTeacher(driver,language)
        # except:
        #     driver.quit()
        #     # raise ValueError(4)
        #     sys.stdout.write("4")
        if status:
            sheet = wb.worksheets[0]
            if sheet.max_row <= 1:
                pass
            else:
                file_path = ''
                file = save_file(file_path)
                wb.save(file)
                wb.close()
                driver.quit()
            # raise ValueError(0)
            sys.stdout.write("0")



def gotoAllTeacher(driver,language):

    time.sleep(0.5)
    lastPage = 1
    try:
        lastPage = driver.find_element(By.XPATH, "//div[@class='pager']//li[last()]").text
    except:
        pass
    i = 1
    while i <= int(lastPage):
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
        numberStudents = len(driver.find_elements(By.XPATH, "//table[@class='people grid']/tbody/tr"))

        j = 1
        while j <= int(numberStudents):
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
            typeOfAgent = driver.find_element(By.XPATH,
                                              "//table[@class='people grid']/tbody/tr[{}]//td[@class='tdName']/p".format(
                                                  j)).text
            if typeOfAgent == 'Учитель':
                driver.find_element(By.XPATH,
                                    f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                massiveParent(driver, 'Учитель',language)
            j += 1
        if i == int(lastPage):
            return True
        i+=1
        driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={i}]").click()

def gotoAllStudentParent(driver,language):
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
    lastPage = 1
    try:
        lastPage = driver.find_element(By.XPATH, "//div[@class='pager']//li[last()]").text
    except:
        pass
    i=1
    while i<=int(lastPage):
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
        numberStudents = len(driver.find_elements(By.XPATH, "//table[@class='people grid']/tbody/tr"))

        j=1

        while j<=int(numberStudents):
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))

            time.sleep(1)
            typeOfAgent = driver.find_element(By.XPATH,
                                              f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdName']/p").text
            time.sleep(1)

            if 'Родитель' in typeOfAgent :
                driver.find_element(By.XPATH,
                                    f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                massiveParent(driver,'Родитель',language)
            j+=1
        if i == int(lastPage):
            return True
        i+=1

        driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={i}]").click()


def massiveGoToAll(driver,parent,language):
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
    lastPage = 1
    try:
        lastPage = driver.find_element(By.XPATH, "//div[@class='pager']//li[last()]").text
    except:
        pass
    i = 1

    while i <= int(lastPage):
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
        numberStudents = len(driver.find_elements(By.XPATH, "//table[@class='people grid']/tbody/tr"))

        j = 1
        while j <= int(numberStudents):
            try:
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
                typeOfAgent = driver.find_element(By.XPATH,
                                                  f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdName']/p").text
                driver.find_element(By.XPATH,
                                    f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                if typeOfAgent == 'Ученик' or typeOfAgent =='Оқушы':
                    massiveStud(driver, parent, language)
                else:
                    massiveParent(driver,typeOfAgent,language)
                j += 1
            except:
                j += 1
                pass

        if i == int(lastPage) :
            return True
        i += 1
        try:
            driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={i}]").click()
        except:
            return True



def gotoAllStudent(driver,parent,language):
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
    lastPage = 1
    try:
        lastPage = driver.find_element(By.XPATH, "//div[@class='pager']//li[last()]").text
    except:
        pass
    i=1

    while i<=int(lastPage):
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
        numberStudents = len(driver.find_elements(By.XPATH, "//table[@class='people grid']/tbody/tr"))

        j=1
        while j<=int(numberStudents):
            try:
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
                typeOfAgent = driver.find_element(By.XPATH, f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdName']/p").text
                if typeOfAgent == "Мұғалім":
                    pass
                elif typeOfAgent == "Ата-ана":
                    pass
                elif typeOfAgent == "Мұғалім, Ата-ана":
                    pass
                elif typeOfAgent == "Учитель, Родитель":
                    pass
                elif typeOfAgent == 'Ученик' or 'Оқушы':
                    driver.find_element(By.XPATH,
                                        f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                    massiveStud(driver,parent,language)
                else:
                    pass
                j+=1
            except:
                j += 1
                pass



        if i == int(lastPage):

            return True
        i+=1
        try:
            driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={i}]").click()
        except:
            return True



def gotoTeacher(driver,language):
    driver.find_element(By.XPATH,
                        "//table[contains(@class,'people')]/tbody//td[@class='tdButtons']//li[@class='iE']").click()
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))

    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))

    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text

    parent = checkParentsExist(studLogin,language)
    if parent:
        writeOnlyParent(parent[0], parent[1], parent[2], parent[3],language)
    else:
        driver.find_element(By.XPATH, "//input[@name='change']").click()
        fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
        tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text

        writeOnlyParent(fullName, studLogin, tempPass, "Учитель",language)
    return True

def goToStudent(driver,parent,language):
    driver.find_element(By.XPATH,"//table[contains(@class,'people')]/tbody/tr[1]/td[@class = 'tdButtons']//a[@title='Редактировать' or @title='Өңдеу']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))

    driver.find_element(By.XPATH,"//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
        driver.find_element(By.XPATH,"//input[@name='change']").click()
    except:
        return False

    fullName = driver.find_element(By.XPATH,"//div[@class= 'owner people_page' ]/div//h2").text
    studLogin = driver.find_element(By.XPATH,"//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    tempPass = driver.find_element(By.XPATH,"//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabReview']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a")))
    classNumber = driver.find_element(By.XPATH,"//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a").text


    #Parents part
    driver.find_element(By.XPATH,"//div[@class ='tabs']//li//a[@id='TabParents']").click()

    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//li/a[@id='buttonAddRelative']")))
    parents = None
    if parent:
        try:
            driver.find_element(By.XPATH,"//div[@class='emptyData']")
        except:
            logins = getOnlyLoginParent(driver,language)
            status = False
            if logins !=0:
                status = checkParentsExist(logins,language)
            if status:
                stepBack(5,driver)
                a = []
                a.append(status)
                create_write_excel(fullName, studLogin, tempPass,language,classNumber,a)
            else:
                parents = goToParent(driver,language)
                create_write_excel(fullName, studLogin, tempPass,language, classNumber, parents)
    else:
        create_write_excel(fullName, studLogin, tempPass,language, classNumber, parents)
    return True

def goToParent(driver,language):
    parents = []
    driver.find_element(By.XPATH, f"//table[contains(@class,'people')]/tbody/tr[2]/td[@class='tdButtons']//li[@class='iE']").click()
    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))
    except:
        stepBack(3,driver)
        return parents
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
    driver.find_element(By.XPATH, "//input[@name='change']").click()
    time.sleep(1)

    fullParentName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page']/div//h2").text
    parentLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    parentPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
    daughterList = []
    daughterList.append(fullParentName)
    daughterList.append(parentLogin)
    daughterList.append(parentPass)
    daughterList.append('Родитель')
    parents.append(daughterList)
    stepBack(8,driver)
    return parents



def gotoDirectParent(driver,language):
    driver.find_element(By.XPATH,
                        "//table[contains(@class,'people')]/tbody//td[@class='tdButtons']//li[@class='iE']").click()

    try:
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))

        driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    except:
        stepBack(1,driver)
        return
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))

    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    parent = checkParentsExist(studLogin,language)

    if parent:
        writeOnlyParent(parent[0],parent[1],parent[2],parent[3],language)

    else:
        driver.find_element(By.XPATH, "//input[@name='change']").click()
        fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
        tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
        writeOnlyParent(fullName,studLogin,tempPass,"Родитель",language)
        stepBack(3,driver)
    return True

def checkParentExistFullName(nameParent,language):
    sheet = wb.worksheets[0]
    foundParent = 0
    for row in sheet.iter_rows(min_col=6, max_col=6):
        for cell in row:
            if str(cell.value) == str(nameParent):
                foundParent = cell
    if foundParent!=0:
        return True
    return False

def massiveStud(driver,parent,language):
    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//a[@id='TabPassword']")))
        driver.find_element(By.XPATH, f"//a[@id='TabPassword']").click()
    except:
        stepBack(1,driver)
        return
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
    driver.find_element(By.XPATH, "//input[@name='change']").click()

    fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabReview']").click()
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a")))
    classNumber = driver.find_element(By.XPATH, "//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a").text

    # Parents part
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabParents']").click()

    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//li/a[@id='buttonAddRelative']")))
    parents = None
    if parent:
        try:
            driver.find_element(By.XPATH, "//div[@class='emptyData']")
            stepBack(5,driver)
        except Exception as err:
            status = False
            logins = getOnlyLoginParent(driver,language)
            if logins ==1:
                return
            if logins != 0:
                status = checkParentsExist(logins,language)

            if status:
                stepBack(5,driver)
                a = []
                a.append(status)
                create_write_excel(fullName, studLogin, tempPass,language, classNumber, a)
                return
            else:
                parents = goToParent(driver,language)
    else:
        stepBack(5,driver)
    create_write_excel(fullName, studLogin, tempPass,language, classNumber, parents)

def getOnlyLoginParent(driver,language):
    driver.find_element(By.XPATH,
                        f"//table[contains(@class,'people')]/tbody/tr[2]/td[@class='tdButtons']//li[@class='iE']").click()
    try:
        WebDriverWait(driver, 3).until(
            EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))
        driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    except:
        stepBack(6,driver)
        return 1
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
    parentLogin = driver.find_element(By.XPATH,
                                      "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    logins = parentLogin
    stepBack(2,driver)
    if len(logins)>0:
        return logins
    return 0

def massiveParent(driver,user,language):
    try:
        driver.find_element(By.XPATH, f"//a[@id='TabReview']").click()
        adminRights = driver.find_element(By.XPATH,"//dt[text()='Админ. права' or  text()='Әкімш. құқықтар']//..//strong[text()='Әкімші' or text()='Администратор']").text
        if adminRights == "Администратор" or "Әкімші":
            stepBack(2,driver)
            return
        else:
            stepBack(1,driver)
    except:
        stepBack(1,driver)

    try:
        driver.find_element(By.XPATH, f"//a[@id='TabPassword']").click()
    except:
        stepBack(1,driver)
        return
    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text

    parent = checkParentsExist(studLogin,language)

    if parent!=False:

        stepBack(2,driver)
        writeOnlyParent(parent[0], parent[1], parent[2], parent[3],language)
    else:

        driver.find_element(By.XPATH, "//input[@name='change']").click()
        fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
        tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
        writeOnlyParent(fullName, studLogin, tempPass,user,language)
        stepBack(3,driver)




def create_write_excel(studName, studLogin, studPass, language,classNumber = -1, parents:list=None):
    role = 'Ученик'


    if language=="kz":
        role = 'Оқушы'

    sheet = wb.worksheets[0]
    last_empty_row = sheet.max_row + 1
    try:
        sheet.cell(row=last_empty_row, column=1).value = role
        sheet.cell(row=last_empty_row, column=2).value = studName
        sheet.cell(row=last_empty_row, column=3).value = studLogin
        sheet.cell(row=last_empty_row, column=4).value = studPass

        if classNumber != -1:
            sheet.cell(row=last_empty_row, column=5).value = classNumber
        if parents:
            parentName = parents[0][1]
            foundParent = 0
            for row in sheet.iter_rows(min_col=3, max_col=3):
                for cell in row:
                    if str(cell.value) == str(parentName):
                        foundParent = cell

            if foundParent != 0:
                row = foundParent.row
                sheet.cell(row=last_empty_row, column=6).value = sheet[row][6].value
                sheet.cell(row=last_empty_row, column=7).value = sheet[row][7].value
                sheet.cell(row=last_empty_row, column=8).value = sheet[row][8].value

            else:

                for i in range(len(parents)):
                    sheet.cell(row=last_empty_row, column=6).value = parents[i][0]
                    sheet.cell(row=last_empty_row, column=7).value = parents[i][1]
                    sheet.cell(row=last_empty_row, column=8).value = parents[i][2]
    except:
        pass



def writeOnlyParent(parentName,parentLogin,parentPass,user,language):
    if language=="kz":
        user = "Мұғалім"
    try:
        sheet = wb.worksheets[0]

        last_empty_row = sheet.max_row + 1
        sheet.cell(row=last_empty_row, column=1).value = user
        sheet.cell(row=last_empty_row, column=2).value = parentName
        sheet.cell(row=last_empty_row, column=3).value = parentLogin
        sheet.cell(row=last_empty_row, column=4).value = parentPass

    except:
        pass

def createExcel(language):
    global wb
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    from openpyxl.styles import Font
    mainFont = Font(size="16")
    mainColour = PatternFill(start_color="ff43e287",
                             end_color="ff43e287",
                             fill_type="solid")
    sheet.column_dimensions['A'].width = 24
    sheet.column_dimensions['B'].width = 24
    sheet.column_dimensions['C'].width = 26
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 22
    sheet.column_dimensions['F'].width = 22
    sheet.column_dimensions['G'].width = 16
    sheet.column_dimensions['H'].width = 24

    sheet.cell(row=1, column=1).font = mainFont
    sheet.cell(row=1, column=2).font = mainFont
    sheet.cell(row=1, column=3).font = mainFont
    sheet.cell(row=1, column=4).font = mainFont
    sheet.cell(row=1, column=5).font = mainFont
    sheet.cell(row=1, column=6).font = mainFont
    sheet.cell(row=1, column=7).font = mainFont
    sheet.cell(row=1, column=8).font = mainFont

    sheet.cell(row=1, column=1).fill = mainColour
    sheet.cell(row=1, column=2).fill = mainColour
    sheet.cell(row=1, column=3).fill = mainColour
    sheet.cell(row=1, column=4).fill = mainColour
    sheet.cell(row=1, column=5).fill = mainColour
    sheet.cell(row=1, column=6).fill = mainColour
    sheet.cell(row=1, column=7).fill = mainColour
    sheet.cell(row=1, column=8).fill = mainColour

    if language == "kz":
        sheet.cell(row=1, column=1).value = "Рөлі"
        sheet.cell(row=1, column=2).value = "Аты-жөні"
        sheet.cell(row=1, column=3).value = "Пайдаланушы аты"
        sheet.cell(row=1, column=4).value = "Құпиясөз"
        sheet.cell(row=1, column=5).value = "Сынып"
        sheet.cell(row=1, column=6).value = "Ата-анасының/заңды өкілінің аты-жөні"
        sheet.cell(row=1, column=7).value = "Пайдаланушы аты"
        sheet.cell(row=1, column=8).value = "Құпиясөз"
    else:
        sheet.cell(row=1, column=1).value = "Роль"
        sheet.cell(row=1, column=2).value = "ФИО пользователя"
        sheet.cell(row=1, column=3).value = "Логин"
        sheet.cell(row=1, column=4).value = "Временный пароль"
        sheet.cell(row=1, column=5).value = "Класс"
        sheet.cell(row=1, column=6).value = "ФИО законного представителя"
        sheet.cell(row=1, column=7).value = "Логин"
        sheet.cell(row=1, column=8).value = "Временный пароль"

def checkParentsExist(parentLogin,language):
    sheet = wb.worksheets[0]
    foundParent = 0
    for row in sheet.iter_rows(min_col=7, max_col=7):
        for cell in row:
            if str(cell.value) == str(parentLogin):
                foundParent = cell


    if foundParent != 0:
        parent = []
        row = foundParent.row
        nameParent = sheet[row][5].value
        loginParent = sheet[row][6].value
        passParent = sheet[row][7].value
        parent.append(nameParent)
        parent.append(loginParent)
        parent.append(passParent)
        return parent
    return False


def stepBack(n,driver):
    for i in range(n):
        driver.back()
        time.sleep(1)

import argparse
parser = argparse.ArgumentParser()
parser.add_argument('--login', type=str, required=True)
parser.add_argument('--password', type=str, required=True)
parser.add_argument('--roles', type=str)
parser.add_argument('--name', type=str,)
parser.add_argument('--groupnumber', type=str)
parser.add_argument('--group', type=str)
parser.add_argument('--chromedriver', type=str)
parser.add_argument('--chromium', type=str)
parser.add_argument('--language', type=str)

args = parser.parse_args()


chrome_options.binary_location = args.chromium
executable_path = args.chromedriver
login = args.login
password = args.password
roles =  args.roles
name = args.name
groupNumber = args.groupnumber
group = args.group
language = args.language
#
main(login,password,roles,name,groupNumber,group,language)

# main("admin.sko","Qwerty1234@",6,"","","True","ru")
# import os



    # create_write_excel("new2",'studLogin','pass',2,[['parentlogin','new1','sdwasdw']])
##1-ученик
# 2-родитель
# 3-ученик и родитель
# 4 teacher
# 5 сотрудники

