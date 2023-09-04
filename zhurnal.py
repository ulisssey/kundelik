from openpyxl import Workbook
from openpyxl.styles import numbers
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from tkinter import Tk, filedialog
from selenium.webdriver.remote.remote_connection import LOGGER
import pyautogui
import sys
import logging
import time


result = '0'

# Конфигурация драйвера 
chrome_options = Options()
# chrome_options.add_argument('--headless')  # Run Chrome in headless mode
chrome_options.add_argument('--no-sandbox')  # Disable sandbox mode to avoid issues with some systems
chrome_options.add_argument('--disable-dev-shm-usage')  # Disable usage of /dev/shm to avoid issues with some systems
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

# Скрипт если пользователь закрыл окно сохранения файла
def save_file(file_path):
    while file_path == '':
        root = Tk()
        root.attributes('-topmost', True)
        root.withdraw()
        file_path = filedialog.asksaveasfilename(parent=root, defaultextension='.xlsx')
        if file_path != '':
            return file_path

# Скрипт для логина и перехода настраницу поиска
def main_code(login, password, driver):
    driver.get("https://login.kundelik.kz/")
    driver.maximize_window()
    # Авторизация в login.kundelik.kz
    try:
        if WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='login']"))):
            driver.find_element(By.XPATH, "//input[@name='login']").send_keys(login)
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
            driver.find_element(By.XPATH, "//input[@type='submit']").click()
            try: 
                if WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='message ']"))):
                    driver.quit()
                    return "1"
            except:
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='header-localization-select__info']/div[1]"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[contains(text(), 'Русский')]"))).click()
                # Путь к странице "Администрирование школы"
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[text()='Образование ']"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@data-test-id='Отчеты']"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@title='Журнал']"))).click()
    except:
        return "4"

def get_data(driver, name, period):
    global wb
    try:
        if (period == 'Все') or (period == 'Барлық'):
            if WebDriverWait(driver, 5).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='emptyData']")):
                return "0"
        else:
            if WebDriverWait(driver, 5).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='emptyData']")):
                driver.quit()
                return "10"
    except:
        if driver.find_elements(By.XPATH, "//div[@class='pager']//li"):
            page_num = driver.find_elements(By.XPATH, "//div[@class='pager']//li")
            num = 1
            pages = int(page_num[-1].text)
            all_changes = []
            while num <= pages+1:
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@class='header-logotype header-logotype_kz']")))
                all = driver.find_elements(By.XPATH, "//table[@class='grid gridLines vam marks ']//tr")
                if not all:
                    pass
                else:
                    if language == 'ru':
                        wb = Workbook()
                        ws = wb.active
                        if (period == 'Итоговые') or (period == 'Қорытындылар'):
                            ws.append(['Дата изменений', 'Время изменений', 'Автор изменений', 'Действие', 'Вид отметки', 'Старое значение', 'Новое значение', 'Ученик', 'Название предмета'])
                        else:
                            ws.append(['Дата изменений', 'Время изменений', 'Автор изменений', 'Действие', 'Старое значение', 'Новое значение', 'Ученик', 'Название предмета', 'Комментарий'])
                        for column_cells in ws.columns:
                            column_letter = column_cells[0].column_letter
                            ws.column_dimensions[column_letter].width = 20
                    else:
                        wb = Workbook()
                        ws = wb.active
                        if (period == 'Итоговые') or (period == 'Қорытындылар'):
                            ws.append(['Өзгеріс күні', 'Өзгеріс уақыты', 'Өзгерістер авторы', 'Әрекет', 'Белгі түрі',  'Ескі мәні', 'Жаңа мән', 'Оқушы', 'Пәннің атауы'])
                        else:
                            ws.append(['Өзгеріс күні', 'Өзгеріс уақыты', 'Өзгерістер авторы', 'Әрекет', 'Ескі мәні', 'Жаңа мән', 'Оқушы', 'Пәннің атауы', 'Комментарий'])
                        for column_cells in ws.columns:
                            column_letter = column_cells[0].column_letter
                            ws.column_dimensions[column_letter].width = 20
                    for id, i in enumerate(all):
                        if ('Автор изменений' in i.text) or ('Өзгерістер авторы' in i.text):
                            continue
                        elif i.find_element(By.TAG_NAME, "td").get_attribute('colspan') == "7":
                            if (period == 'Итоговые') or (period == 'Қорытындылар'):
                                date = i.text
                            else:
                                if len(name) != 0:
                                    if all[id-1].get_attribute('class') == i.get_attribute('class'):
                                        last = name.split(" ")[0].lower()
                                        firstname = name.split(" ")[1][:1].lower()
                                        name = f"{last} {firstname}"
                                        if name not in all[id-1].find_elements(By.TAG_NAME, 'td')[8].text.lower():
                                            continue
                                        else:
                                            all_changes[-1].append(i.text)
                                else:
                                    if (all_changes[-1][1] == all[id-1].find_elements(By.TAG_NAME, 'td')[0].text) and (all_changes[-1][2] == all[id-1].find_elements(By.TAG_NAME, 'td')[1].text) and (all_changes[-1][3] == all[id-1].find_elements(By.TAG_NAME, 'td')[3].text) and (all_changes[-1][4] == all[id-1].find_elements(By.TAG_NAME, 'td')[6].text) and (all_changes[-1][5] == all[id-1].find_elements(By.TAG_NAME, 'td')[7].text) and (all_changes[-1][6] == all[id-1].find_elements(By.TAG_NAME, 'td')[8].text):
                                        all_changes[-1].append(i.text)
                        elif (i.find_element(By.TAG_NAME, "td").get_attribute('colspan') == '10') or (i.find_element(By.TAG_NAME, "td").get_attribute('colspan') == '9'):
                            date = i.text
                            if (period == 'Итоговые') or (period == 'Қорытындылар'):
                                continue
                        else:
                            if (typeOfAction == 'Добавление' and i.get_attribute("class") == 'create') or (typeOfAction == 'Удаление' and i.get_attribute("class") == 'delete') or (typeOfAction == 'Изменение' and i.get_attribute("class") == 'update') or len(typeOfAction) == 0 or (typeOfAction == 'Все') or (typeOfAction == 'Барлық') or (typeOfAction == 'Қосу' and i.get_attribute("class") == 'create') or (typeOfAction == 'Жою' and i.get_attribute("class") == 'delete') or (typeOfAction == 'Өзгерту' and i.get_attribute("class") == 'update'):
                                if len(name) != 0:
                                    try:
                                        last = name.split(" ")[0].lower()
                                        firstname = name.split(" ")[1][:1].lower()
                                        name = f"{last} {firstname}"
                                        if (period == 'Итоговые') or (period == 'Қорытындылар'):
                                            if name not in i.find_elements(By.TAG_NAME, 'td')[6].text.lower():
                                                continue
                                        else:
                                            if name not in i.find_elements(By.TAG_NAME, 'td')[8].text.lower():
                                                continue
                                    except:
                                        pass
                                try:
                                    if (period == 'Итоговые') or (period == 'Қорытындылар'):
                                        time = i.find_elements(By.TAG_NAME, 'td')[0].text
                                        author = i.find_elements(By.TAG_NAME, 'td')[1].text 
                                        action = i.find_elements(By.TAG_NAME, 'td')[2].text
                                        grade_type = i.find_elements(By.TAG_NAME, 'td')[3].text
                                        old_value = i.find_elements(By.TAG_NAME, 'td')[4].text
                                        new_value = i.find_elements(By.TAG_NAME, 'td')[5].text
                                        student = i.find_elements(By.TAG_NAME, 'td')[6].text
                                        subject = subjects
                                        change = [date, time, author, action, grade_type, old_value, new_value, student, subject]
                                        all_changes.append(change)
                                    else:
                                        time = i.find_elements(By.TAG_NAME, 'td')[0].text
                                        author = i.find_elements(By.TAG_NAME, 'td')[1].text 
                                        action = i.find_elements(By.TAG_NAME, 'td')[3].text
                                        old_value = i.find_elements(By.TAG_NAME, 'td')[6].text
                                        new_value = i.find_elements(By.TAG_NAME, 'td')[7].text
                                        student = i.find_elements(By.TAG_NAME, 'td')[8].text
                                        if len(subjects) != 0:
                                            subject = subjects
                                        else:
                                            subject = i.find_elements(By.TAG_NAME, 'td')[9].text
                                        change = [date, time, author, action, old_value, new_value, student, subject]
                                        all_changes.append(change)
                                except:
                                    continue
                num += 1
                if num == pages+1:
                    break
                else:
                    driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={num}]").click()
            if all_changes == []:
                return "9"
            for change in all_changes:
                ws.append(change)
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
            return "0"
        else:
            all_changes = []
            all = driver.find_elements(By.XPATH, "//table[@class='grid gridLines vam marks ']//tr")
            if not all:
                return "10"
            else:
                if language == 'ru':
                    wb = Workbook()
                    ws = wb.active
                    if (period == 'Итоговые') or (period == 'Қорытындылар'):
                        ws.append(['Дата изменений', 'Время изменений', 'Автор изменений', 'Действие', 'Вид отметки', 'Старое значение', 'Новое значение', 'Ученик', 'Название предмета'])
                    else:
                        ws.append(['Дата изменений', 'Время изменений', 'Автор изменений', 'Действие', 'Старое значение', 'Новое значение', 'Ученик', 'Название предмета', 'Комментарий'])
                    for column_cells in ws.columns:
                        column_letter = column_cells[0].column_letter
                        ws.column_dimensions[column_letter].width = 20
                else:
                    wb = Workbook()
                    ws = wb.active
                    if (period == 'Итоговые') or (period == 'Қорытындылар'):
                        ws.append(['Өзгеріс күні', 'Өзгеріс уақыты', 'Өзгерістер авторы', 'Әрекет', 'Белгі түрі',  'Ескі мәні', 'Жаңа мән', 'Оқушы', 'Пәннің атауы'])
                    else:
                        ws.append(['Өзгеріс күні', 'Өзгеріс уақыты', 'Өзгерістер авторы', 'Әрекет', 'Ескі мәні', 'Жаңа мән', 'Оқушы', 'Пәннің атауы', 'Комментарий'])
                    for column_cells in ws.columns:
                        column_letter = column_cells[0].column_letter
                        ws.column_dimensions[column_letter].width = 20
                for id, i in enumerate(all):
                    if ('Автор изменений' in i.text) or ('Өзгерістер авторы' in i.text):
                        continue
                    elif i.find_element(By.TAG_NAME, "td").get_attribute('colspan') == "7":
                        if (period == 'Итоговые') or (period == 'Қорытындылар'):
                            date = i.text
                        else:
                            if len(name) != 0:
                                if all[id-1].get_attribute('class') == i.get_attribute('class'):
                                    last = name.split(" ")[0].lower()
                                    firstname = name.split(" ")[1][:1].lower()
                                    name = f"{last} {firstname}"
                                    if name not in all[id-1].find_elements(By.TAG_NAME, 'td')[8].text.lower():
                                        continue
                                    else:
                                        all_changes[-1].append(i.text)
                            else:
                                if (all_changes[-1][1] == all[id-1].find_elements(By.TAG_NAME, 'td')[0].text) and (all_changes[-1][2] == all[id-1].find_elements(By.TAG_NAME, 'td')[1].text) and (all_changes[-1][3] == all[id-1].find_elements(By.TAG_NAME, 'td')[3].text) and (all_changes[-1][4] == all[id-1].find_elements(By.TAG_NAME, 'td')[6].text) and (all_changes[-1][5] == all[id-1].find_elements(By.TAG_NAME, 'td')[7].text) and (all_changes[-1][6] == all[id-1].find_elements(By.TAG_NAME, 'td')[8].text):
                                    all_changes[-1].append(i.text)
                    elif (i.find_element(By.TAG_NAME, "td").get_attribute('colspan') == '10') or (i.find_element(By.TAG_NAME, "td").get_attribute('colspan') == '9'):
                        date = i.text
                        if (period == 'Итоговые') or (period == 'Қорытындылар'):
                            continue
                    else:
                        if (typeOfAction == 'Добавление' and i.get_attribute("class") == 'create') or (typeOfAction == 'Удаление' and i.get_attribute("class") == 'delete') or (typeOfAction == 'Изменение' and i.get_attribute("class") == 'update') or len(typeOfAction) == 0 or (typeOfAction == 'Все') or (typeOfAction == 'Барлық') or (typeOfAction == 'Қосу' and i.get_attribute("class") == 'create') or (typeOfAction == 'Жою' and i.get_attribute("class") == 'delete') or (typeOfAction == 'Өзгерту' and i.get_attribute("class") == 'update'):
                            if len(name) != 0:
                                try:
                                    last = name.split(" ")[0].lower()
                                    firstname = name.split(" ")[1][:1].lower()
                                    name = f"{last} {firstname}"
                                    if (period == 'Итоговые') or (period == 'Қорытындылар'):
                                        if name not in i.find_elements(By.TAG_NAME, 'td')[6].text.lower():
                                            continue
                                    else:
                                        if name not in i.find_elements(By.TAG_NAME, 'td')[8].text.lower():
                                            continue
                                except:
                                    pass
                            try:
                                if (period == 'Итоговые') or (period == 'Қорытындылар'):
                                    time = i.find_elements(By.TAG_NAME, 'td')[0].text
                                    author = i.find_elements(By.TAG_NAME, 'td')[1].text 
                                    action = i.find_elements(By.TAG_NAME, 'td')[2].text
                                    grade_type = i.find_elements(By.TAG_NAME, 'td')[3].text
                                    old_value = i.find_elements(By.TAG_NAME, 'td')[4].text
                                    new_value = i.find_elements(By.TAG_NAME, 'td')[5].text
                                    student = i.find_elements(By.TAG_NAME, 'td')[6].text
                                    subject = subjects
                                    change = [date, time, author, action, grade_type, old_value, new_value, student, subject]
                                    all_changes.append(change)
                                else:
                                    time = i.find_elements(By.TAG_NAME, 'td')[0].text
                                    author = i.find_elements(By.TAG_NAME, 'td')[1].text 
                                    action = i.find_elements(By.TAG_NAME, 'td')[3].text
                                    old_value = i.find_elements(By.TAG_NAME, 'td')[6].text
                                    new_value = i.find_elements(By.TAG_NAME, 'td')[7].text
                                    student = i.find_elements(By.TAG_NAME, 'td')[8].text
                                    if len(subjects) != 0:
                                        subject = subjects
                                    else:
                                        subject = i.find_elements(By.TAG_NAME, 'td')[9].text
                                    change = [date, time, author, action, old_value, new_value, student, subject]
                                    all_changes.append(change)
                            except:
                                continue
                if all_changes == []:
                    return "10"
                for change in all_changes:
                    ws.append(change)
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
    return "0"

def search(login, password, studyYear, classNumber, period, name, typeOfAction, subjects, dateFrom, dateTo, massiveUnloading, language):
    global subject
    
    LOGGER.setLevel(logging.WARNING)
    driver = webdriver.Chrome(options=chrome_options, service=Service(executable_path=executable_path))
    result = main_code(login, password, driver)
    if result == "1":
        driver.quit()
        return "1"
    elif result == "4":
        driver.quit()
        return result
    try:
        WebDriverWait(driver, 4).until(lambda driver: driver.find_element(By.XPATH, f"//a[text()='{studyYear}']")).click()
    except:
        driver.quit()
        return "12"
    try:
        WebDriverWait(driver, 4).until(lambda driver: driver.find_element(By.XPATH, f"//a[text()='{classNumber}']")).click()
    except:
        driver.quit()
        return "20"
    try:
        if len(subjects) != 0:
            select = Select(driver.find_element(By.ID, 'subject'))
            select.select_by_visible_text(subjects)
    except:
        driver.quit()
        return "15"
    if len(period) != 0:
        if (period != 'Все') and (period != 'Барлық'):
            if language == 'kz':
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='header-localization-select__info']/div[1]"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[contains(text(), 'Қазақ')]"))).click()            
            driver.find_element(By.XPATH, f"//a[text()='{period}']").click()
            if len(dateFrom) != 0:
                driver.find_element(By.XPATH, "//div[@class='header-localization-select__info']/div[1]").click()
                driver.find_element(By.XPATH, "//a[contains(text(), 'Русский')]").click()
                WebDriverWait(driver, 10).until(lambda driver: driver.find_element(By.XPATH, "//input[@id='datefrom']")).click()
                WebDriverWait(driver, 10).until(lambda driver: driver.find_element(By.XPATH, "//div[@id='calendar']"))
                if dateFrom.split('.')[1] == '01':
                    if 'Январь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='янв']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '02':
                    if 'Февраль' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='фев']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '03':
                    if 'Март' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='мар']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '04':
                    if 'Апрель' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='апр']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '05':
                    if 'Май' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='май']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '06':
                    if 'Июнь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='июн']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '07':
                    if 'Июль' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='июл']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '08':
                    if 'Август' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='авг']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '09':
                    if 'Сентябрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='сен']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '10':
                    if 'Октябрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='окт']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '11':
                    if 'Ноябрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='ноя']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                elif dateFrom.split('.')[1] == '12':
                    if 'Декабрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='дек']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateFrom.split('.')[0]}]").click()
            if len(dateTo) != 0:
                WebDriverWait(driver, 10).until(lambda driver: driver.find_element(By.XPATH, "//input[@id='dateto']")).click()
                WebDriverWait(driver, 10).until(lambda driver: driver.find_element(By.XPATH, "//div[@id='calendar']"))
                if dateTo.split('.')[1] == '01':
                    if 'Январь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='янв']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '02':
                    if 'Февраль' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='фев']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '03':
                    if 'Март' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='мар']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '04':
                    if 'Апрель' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='апр']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '05':
                    if 'Май' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='май']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '06':
                    if 'Июнь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='июн']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '07':
                    if 'Июль' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='июл']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '08':
                    if 'Август' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='авг']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '09':
                    if 'Сентябрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='сен']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        time.sleep(1)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '10':
                    if 'Октябрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='окт']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '11':
                    if 'Ноябрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='ноя']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                elif dateTo.split('.')[1] == '12':
                    if 'Декабрь' not in driver.find_element(By.XPATH, "//div[@class='current']").text:
                        driver.find_element(By.XPATH, "//div[@class='current']/a").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='year']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, "//a[text()='дек']").click()
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                    else:
                        WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//div[@class='content']"))
                        time.sleep(2)
                        driver.find_element(By.XPATH, f"//a[text()={dateTo.split('.')[0]}]").click()
                driver.find_element(By.XPATH, "//input[@id='button']").click()
                if language == 'kz':
                    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='header-localization-select__info']/div[1]"))).click()
                    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[contains(text(), 'Қазақ')]"))).click()
            result = get_data(driver, name, period)
            if result != "0":
                return result
            if wb.active.max_row <= 1:
                pass
            else:
                file_path = ''
                file = save_file(file_path)
                wb.save(file)
                wb.close()
            driver.quit()
            return result
        else:
            if language == 'kz':
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='header-localization-select__info']/div[1]"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[contains(text(), 'Қазақ')]"))).click()
            for i in range(1, 5):
                driver.find_element(By.XPATH, f"//a[text()='{i}-я четверть' or text()='{i}-тоқсан']").click()
                result = get_data(driver, name, period)
            if result != "0":
                return result
            if wb.active.max_row <= 1:
                pass
            else:
                file_path = ''
                file = save_file(file_path)
                wb.save(file)
                wb.close()
            driver.quit()
            return result

    
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('--login', type=str, required=True)
parser.add_argument('--password', type=str, required=True)
parser.add_argument('--name', type=str, nargs='?')
parser.add_argument('--studyYear', type=str, required=True)
parser.add_argument('--classNumber', type=str, required=True)
parser.add_argument('--period', type=str, nargs='?')
parser.add_argument('--typeOfAction', type=str, nargs='?')
parser.add_argument('--subjects', type=str, nargs='?')
parser.add_argument('--dateFrom', type=str, nargs='?')
parser.add_argument('--dateTo', type=str, nargs='?')
parser.add_argument('--massiveUnloading', type=str, nargs='?')
parser.add_argument('--language', type=str, required=True)
parser.add_argument('--chromedriver', type=str, required=True)
parser.add_argument('--chromium', type=str, required=True)
args = parser.parse_args()
executable_path = rf"{args.chromedriver}"
chrome_options.binary_location = rf"{args.chromium}"
name = args.name
if args.period == '1-і тоқсан':
    period = '1-тоқсан'
elif args.period == '2-і тоқсан':
    period = '2-тоқсан'
elif args.period == '3-і тоқсан':
    period = '3-тоқсан'
elif args.period == '4-і тоқсан':
    period = '4-тоқсан'
else:
    period = args.period
typeOfAction = args.typeOfAction
subjects = args.subjects
dateFrom = args.dateFrom
dateTo = args.dateTo
massiveUnloading = args.massiveUnloading
language = args.language
try:
    res = search(args.login, args.password, args.studyYear, args.classNumber, period, name, typeOfAction, subjects, args.dateFrom, args.dateTo, massiveUnloading, language)
    sys.stdout.write(res)
except:
    sys.stdout.write("4")