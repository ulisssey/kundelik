from openpyxl import Workbook
from openpyxl.styles import numbers
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from datetime import datetime
from tkinter import Tk, filedialog
import time
import re
import sys
import logging
from selenium.webdriver.remote.remote_connection import LOGGER


school = ''
result = '0'

# Засечение времени
day = datetime.now().day
month = datetime.now().month
year = datetime.now().year

# Конфигурация драйвера 
chrome_options = Options()
# chrome_options.add_argument('--headless')  # Run Chrome in headless mode
chrome_options.add_argument('--no-sandbox')  # Disable sandbox mode to avoid issues with some systems
chrome_options.add_argument('--disable-dev-shm-usage')  # Disable usage of /dev/shm to avoid issues with some systems
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

# Скрипт если пользователь закрыл окно сохранения файла
def save_file(file_path):
    while file_path == '':
        root = Tk()
        root.attributes('-topmost', True)
        root.withdraw()
        file_path = filedialog.asksaveasfilename(parent=root, defaultextension='.xlsx')
        if file_path != '':
            return file_path

# Скрипт для логина и перехода настраницу поиска
def main_code(login, password, driver):
    driver.get("https://login.kundelik.kz/")
    driver.maximize_window()
    # Авторизация в login.kundelik.kz
    try:
        if WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='login']"))):
            driver.find_element(By.XPATH, "//input[@name='login']").send_keys(login)
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
            driver.find_element(By.XPATH, "//input[@type='submit']").click()
            try: 
                if WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='message ']"))):
                    driver.quit()
                    return "1"
            except:
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='header-localization-select__info']/div[1]"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[contains(text(), 'Русский')]"))).click()
                status = get_school(driver)
                if status == "4":
                    return "4"
                # Путь к странице "Администрирование школы"
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[text()='Образование ']"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[text()='Моя школа ']"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@title='Администрирование школы']"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[text()='Список людей']"))).click()
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//h2[text()='Поиск людей']")))
    except:
        pass
    
# Скрипт для получения названия школ сотрудникам
def get_school(driver):
    try:
        driver.find_element(By.XPATH, "//a[text()='Профиль ']").click()
        school = WebDriverWait(driver, 60).until(lambda driver: driver.find_element(By.XPATH, "//dt[text()='Школа:']/..//a")).text
    except:
        return "4" 
# Скрипт для получения данных из страницы ученика
def get_data(driver, lang):
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//dl[@class='info s2']")))
    all_text = driver.find_elements(By.XPATH, "//dl[@class='info s2']/dd")
    fio = all_text[0].text
    gender = all_text[1].text
    if lang == 'kz':
        if gender == 'Мужской':
            gender = 'Ер'
        else:
            gender = 'Әйел'
    birth_date = all_text[3].text
    email = all_text[4].text
    all_text1 = driver.find_elements(By.XPATH, "//dl[@class='info big']/dd")
    school_class = all_text1[0].text
    match = re.search(r'\d+', school_class)
    parallel = int(match.group())
    # Личные данные
    driver.find_element(By.XPATH, "//a[@id='TabPersonal']").click()
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//label[text()='Фамилия']")))
    iin = driver.find_element(By.XPATH, "//input[@id='personalNumber']").get_attribute('value')
    select = Select(driver.find_element(By.XPATH, "//select[@id='nationality']"))
    nationality = select.first_selected_option.text
    if nationality == '- Выберите национальность -':
        nationality = None
    if lang == 'kz':
        if nationality == 'казах / казашка':
            nationality = 'қазақ'
        elif nationality == 'русский / русская':
            nationality = 'орыс'
        elif nationality == 'немец / немка':
            nationality == 'неміс'
        elif nationality == 'азербайджанец / азербайджанка':
            nationality == 'әзірбайжан'
        elif nationality == 'башкир / башкирка':
            nationality == 'башқұрт'
        elif nationality == 'белорус / белоруска':
            nationality == 'белорус'
        elif nationality == 'грек / гречанка':
            nationality == 'грек'
        elif nationality == 'дунган / дунганка':
            nationality == 'дунган'
        elif nationality == 'еврей / еврейка':
            nationality == 'еврей'
        elif nationality == 'ингуш / ингушка':
            nationality == 'ингуш'
        elif nationality == 'кореец / кореянка':
            nationality == 'кәріс'
        elif nationality == 'курд / курдка':
            nationality == 'күрт'
        elif nationality == 'кыргыз / кыргызка':
            nationality == 'қырғыз'
        elif nationality == 'молдованин / молодованка':
            nationality == 'молдаван'
        elif nationality == 'поляк / полька':
            nationality == 'поляк'
        elif nationality == 'таджик / таджичка':
            nationality == 'тәжік'
        elif nationality == 'татарин / татарка':
            nationality == 'татар'
        elif nationality == 'турок / турчанка':
            nationality == 'түрік'
        elif nationality == 'узбек / узбечка':
            nationality == 'өзбек'
        elif nationality == 'уйгур / уйгурка':
            nationality == 'ұйғыр'
        elif nationality == 'украинец / украинка':
            nationality == 'украин'
        elif nationality == 'чеченец / чеченка':
            nationality == 'шешен'
        elif nationality == 'чуваш / чувашка':
            nationality == 'чуваш'
        elif nationality == 'другие национальности':
            nationality == 'басқа ұлттар'
    select = Select(driver.find_element(By.XPATH, "//select[@id='nativeLanguage']"))
    native_lg = select.first_selected_option.text
    if native_lg == '- Выберите язык -':
        native_lg = None
    if lang == 'kz':
        if native_lg == 'Казахский':
            native_lg == 'Қазақша'
        elif native_lg == 'Русский':
            native_lg = 'Орысша'
        elif native_lg == 'Азербайджанский':
            native_lg = 'Әзірбайжан тілі'
        elif native_lg == 'Башкирский':
            native_lg = 'Башқұрт тілі'
        elif native_lg == 'Белорусский':
            native_lg = 'Белорусс тілі'
        elif native_lg == 'Греческий':
            native_lg = 'Грек тілі'
        elif native_lg == 'Дунганский':
            native_lg = 'Дүнген тілі'
        elif native_lg == 'Иврит':
            native_lg = 'Иврит тілі'
        elif native_lg == 'Ингушский':
            native_lg = 'Ингуш тілі'
        elif native_lg == 'Корейский':
            native_lg = 'Корей тілі'
        elif native_lg == 'Курдский':
            native_lg = 'Күрд тілі'
        elif native_lg == 'Кыргызский':
            native_lg = 'Қырғыз тілі'
        elif native_lg == 'Молдавский':
            native_lg = 'Молдав тілі'
        elif native_lg == 'Мордовский':
            native_lg = 'Мордов тілі'
        elif native_lg == 'Немецкий':
            native_lg = 'Неміс тілі'
        elif native_lg == 'Польский':
            native_lg = 'Поляк тілі'
        elif native_lg == 'Таджикский':
            native_lg = 'Тәжік тілі'
        elif native_lg == 'Татарский':
            native_lg = 'Татар тілі'
        elif native_lg == 'Турецкий':
            native_lg = 'Түрік тілі'
        elif native_lg == 'Узбекский':
            native_lg = 'Өзбек тілі'
        elif native_lg == 'Уйгурский':
            native_lg = 'Ұйғыр тілі'
        elif native_lg == 'Украинский':
            native_lg = 'Украин тілі'
        elif native_lg == 'Чеченский':
            native_lg = 'Шешен тілі'
        elif native_lg == 'Чувашский':
            native_lg = 'Чуваш тілі'
        elif native_lg == 'Другие языки':
            native_lg = 'Басқа тілдер'
    svidetelstvo = driver.find_element(By.XPATH, "//input[@id='bcert_docnum']").get_attribute('value')
    kem_vydan = driver.find_element(By.XPATH, "//input[@id='bcert_issby']").get_attribute('value')
    data_vydachy = driver.find_element(By.XPATH, "//input[@id='birthcertificatedate']").get_attribute('value')
    mesto_vydachy = driver.find_element(By.XPATH, "//input[@id='bcert_place']").get_attribute('value')
    select = Select(driver.find_element(By.XPATH, "//select[@id='medid']"))
    medid = select.first_selected_option.text
    select = Select(driver.find_element(By.XPATH, "//select[@id='physid']"))
    physid = select.first_selected_option.text
    address = driver.find_element(By.XPATH, "//input[@id='UserControlPersonEdit_Tabs_UserControlPersonEditPersonal_ActualAddress']").get_attribute('value')
    mphone = driver.find_element(By.XPATH, "//input[@name='mphone']").get_attribute('value')
    hphone = driver.find_element(By.XPATH, "//input[@name='hphone']").get_attribute('value')
    # Миграция
    driver.find_element(By.XPATH, "//a[@id='TabMigration']").click()
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='wasInPreschoolInstitution']")))
    pre_school = driver.find_element(By.XPATH, "//input[@id='wasInPreschoolInstitution']").get_attribute('checked') 
    if pre_school == 'checked':
        pre_school = 'Да'
    else:
        pre_school = 'Нет'
    # Родственники
    driver.find_element(By.XPATH, "//a[@id='TabParents']").click()
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@id='buttonAddRelative']")))
    try:
        if driver.find_element(By.XPATH, "//p[text()='У пользователя не настроены родственные связи']"):
            fio_parent = ''
            work_place = ''
            dolzhnost = ''
            parent_iin = ''
            parent_phone = ''
    except:
        driver.find_element(By.XPATH, "//a[@title='Редактировать']").click()
        all_text = driver.find_elements(By.XPATH, "//dl[@class='info s2']/dd")
        fio_parent = all_text[0].text
        all_text1 = driver.find_elements(By.XPATH, "//dl[@class='info big']/dt")
        all_text2 = driver.find_elements(By.XPATH, "//dl[@class='info big']/dd")
        if all_text1[0].text == 'Должность':
            work_place = school
            dolzhnost = all_text2[0].text
            driver.find_element(By.XPATH, "//a[@id='TabPersonal']").click()
            WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='personalNumber']")))
            parent_iin = driver.find_element(By.XPATH, "//input[@id='personalNumber']").get_attribute('value')
            parent_phone = driver.find_element(By.XPATH, "//input[@name='mphone']").get_attribute('value')
        else:
            driver.find_element(By.XPATH, "//a[@id='TabPersonal']").click()
            WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='personalNumber']")))
            parent_iin = driver.find_element(By.XPATH, "//input[@id='personalNumber']").get_attribute('value')
            parent_phone = driver.find_element(By.XPATH, "//input[@name='mphone']").get_attribute('value')
            work_place = driver.find_element(By.XPATH, "//textarea[@id='workplace']").get_attribute('value')
            dolzhnost = driver.find_element(By.XPATH, "//input[@id='workPosition']").get_attribute('value')
        driver.back()
        driver.back()
    return [fio, gender, birth_date, iin, school_class, parallel, mphone, hphone, address, email, nationality, native_lg, svidetelstvo, data_vydachy, kem_vydan, mesto_vydachy, medid, physid, pre_school, fio_parent, work_place, dolzhnost, parent_iin, parent_phone]

# Скрипт для массовой выгрузки
def get_all(login, password, lang):
    pupil = []
    LOGGER.setLevel(logging.WARNING)
    driver = webdriver.Chrome(options=chrome_options, service=Service(executable_path=executable_path))
    if lang == 'ru':
        wb = Workbook()
        # wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
        # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
        ws = wb.active
        ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
        col = ws.column_dimensions['D']
        col1 = ws.column_dimensions['V']
        col1.number_format = numbers.FORMAT_TEXT
        col.number_format = numbers.FORMAT_TEXT
        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter
            ws.column_dimensions[column_letter].width = 20
            # wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
    else:
        wb = Workbook()
        # wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
        # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
        ws = wb.active
        ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
        col = ws.column_dimensions['D']
        col1 = ws.column_dimensions['V']
        col1.number_format = numbers.FORMAT_TEXT
        col.number_format = numbers.FORMAT_TEXT
        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter
            ws.column_dimensions[column_letter].width = 20
            # wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
    result = main_code(login, password, driver)
    if result == "1":
        return result
    elif result == "4":
        return result
    page_num = driver.find_elements(By.XPATH, "//div[@class='pager']//li")
    num = 1
    pages = int(page_num[-1].text)
    try:
        while num <= pages+1:
            WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@class='header-logotype header-logotype_kz']")))
            all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
            if not all:
                pass
            else:
                children = []
                for i in all:
                    children.append(i.text)
                for j in children:
                    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, f"//a[text()='{j}']")))
                    driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
                    rows = get_data(driver, lang)
                    if rows == None:
                        driver.back()
                    else:
                        pupil.append(rows)
                        # wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
                        driver.back()
                        driver.back()
                        driver.back()
                        driver.back()
            num += 1
            if num == pages+1:
                break
            else:
                driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={num}]").click()
        for p in pupil:
            ws.append(p)
            col = ws.column_dimensions['D']
            col1 = ws.column_dimensions['V']
            col1.number_format = numbers.FORMAT_TEXT
            col.number_format = numbers.FORMAT_TEXT
        file_path = ''
        file = save_file(file_path)
        wb.save(file)
        return "0"
    except:
        return "4"

# Скрипт для выгрузки по ФИО
def search_by_fio(login, password, fio, lang):
    try:
        LOGGER.setLevel(logging.WARNING)
        driver = webdriver.Chrome(options=chrome_options, service=Service(executable_path=executable_path))
        result = main_code(login, password, driver)
        if result == "1":
            return result
        elif result == "4":
            return result
        driver.find_element(By.XPATH, "//input[@id='search']").send_keys(fio)
        time.sleep(0.8)
        driver.find_element(By.XPATH, "//input[@id='go']").click()
        if driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']"):
            if lang == 'ru':
                wb = Workbook()
                # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
                # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
                ws = wb.active
                ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                for column_cells in ws.columns:
                    column_letter = column_cells[0].column_letter
                    ws.column_dimensions[column_letter].width = 20
                    # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
            else:
                wb = Workbook()
                # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
                # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
                ws = wb.active
                ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                for column_cells in ws.columns:
                    column_letter = column_cells[0].column_letter
                    ws.column_dimensions[column_letter].width = 20
                    # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
            all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
            children = []
            for i in all:
                children.append(i.text)
            for j in children:
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, f"//a[text()='{j}']")))
                driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
                rows = get_data(driver, lang)
                if rows == None:
                    driver.back()
                else:
                    ws.append(rows)
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
                    driver.back()
                    driver.back()
                    driver.back()
                    driver.back()
            # root = Tk()
            # root.withdraw()
            # file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
            file_path = ''
            file = save_file(file_path)
            wb.save(file)
            return "0"
        else:
            if lang == 'ru':
            #     wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
            #     ws = wb.active
            #     ws.append([fio] + ['']*23 + ['Не найден ученик с таким ФИО'])
            #     wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
                return "9"
            else:
                # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
                # ws = wb.active
                # ws.append([fio] + ['']*23 + ['Мұндай аты-жөнімен оқушы табылмады'])
                # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
                return "9"
    except:
        return "4"
# Скрипт для выгрузки по классам
def search_by_class(login, password, class_text, lang):
    try:
        pupil = []
        LOGGER.setLevel(logging.WARNING)
        driver = webdriver.Chrome(options=chrome_options, service=Service(executable_path=executable_path))
        result = main_code(login, password, driver)
        if result == "1":
            return "1"
        elif result == "4":
            return result
        driver.find_element(By.XPATH, "//input[@id='class']").send_keys(class_text)
        time.sleep(0.8)
        driver.find_element(By.XPATH, "//input[@id='go']").click()
        time.sleep(5)
        if driver.find_elements(By.XPATH, "//div[@class='pager']//li"):
            page_num = driver.find_elements(By.XPATH, "//div[@class='pager']//li")
            num = 1
            pages = int(page_num[-1].text)
            all_pupil = []
            while num <= pages+1:
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@class='header-logotype header-logotype_kz']")))
                all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
                if not all:
                    pass
                else:
                    if lang == 'ru':
                        wb = Workbook()
                        # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                        # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                        ws = wb.active
                        ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
                        for column_cells in ws.columns:
                            column_letter = column_cells[0].column_letter
                            ws.column_dimensions[column_letter].width = 20
                        # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                    else:
                        wb = Workbook()
                        # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                        # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                        ws = wb.active
                        ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
                        for column_cells in ws.columns:
                            column_letter = column_cells[0].column_letter
                            ws.column_dimensions[column_letter].width = 20
                        # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                    for i in all:
                        all_pupil.append(i.text)
                    children = []
                    for i in all:
                        children.append(i.text)
                    for j in children:
                        WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, f"//a[text()='{j}']")))
                        driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
                        rows = get_data(driver, lang)
                        if rows == None:
                            driver.back()
                        else:
                            pupil.append(rows)
                            
                            # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                            driver.back()
                            driver.back()
                            driver.back()
                            driver.back()
                num += 1
                if num == pages+1:
                    break
                else:
                    driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={num}]").click()
            for p in pupil:
                ws.append(p)
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
            file_path = ''
            file = save_file(file_path)
            wb.save(file)

            if len(all_pupil) == 0:
                if lang == 'kz':
                    return "20"
                else:
                    return "20"
            else:
                all_pupil.clear()
            return "0"
        else:
            all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
            if not all:
                return "20"
            else:
                if lang == 'ru':
                    wb = Workbook()
                    # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                    # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                    ws = wb.active
                    ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
                    for column_cells in ws.columns:
                        column_letter = column_cells[0].column_letter
                        ws.column_dimensions[column_letter].width = 20
                    # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                else:
                    wb = Workbook()
                    # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                    # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                    ws = wb.active
                    ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
                    for column_cells in ws.columns:
                        column_letter = column_cells[0].column_letter
                        ws.column_dimensions[column_letter].width = 20
                    # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                children = []
                for i in all:
                    children.append(i.text)
                for j in children:
                    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, f"//a[text()='{j}']")))
                    driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
                    rows = get_data(driver, lang)
                    if rows == None:
                        driver.back()
                    else:
                        pupil.append(rows)
                        # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                        driver.back()
                        driver.back()
                        driver.back()
                        driver.back()
            for p in pupil:
                ws.append(p)
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT            
            file_path = ''
            file = save_file(file_path)
            wb.save(file)
            return "0"
    except:
        return "4"
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('--login', type=str, required=True)
parser.add_argument('--password', type=str, required=True)
parser.add_argument('--lang', type=str, required=True)
parser.add_argument('--fio', type=str)
parser.add_argument('--clas', type=str)
parser.add_argument('--mass', type=str)
parser.add_argument('--kundelik_id', type=str)
parser.add_argument('--accessToken', type=str)
parser.add_argument('--chromedriver', type=str)
parser.add_argument('--chromium', type=str)
args = parser.parse_args()
executable_path = rf"{args.chromedriver}"
chrome_options.binary_location = rf"{args.chromium}"
if args.fio != '':
    res = search_by_fio(args.login, args.password, args.fio, args.lang)
    sys.stdout.write(res)
if args.clas != '':
    res = search_by_class(args.login, args.password, args.clas, args.lang)
    sys.stdout.write(res)
if args.mass == 'true':
    res = get_all(args.login, args.password, args.lang)
    sys.stdout.write(res)